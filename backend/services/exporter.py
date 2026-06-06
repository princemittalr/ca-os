import datetime
from io import BytesIO
from typing import Dict, List, Any

# openpyxl imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# reportlab imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

# Excel Styling Constants (Navy Theme)
HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SUMMARY_TITLE_FILL = PatternFill(start_color="0F2537", end_color="0F2537", fill_type="solid")
SUMMARY_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")

KPI_TITLE_FONT = Font(name="Calibri", size=9, bold=True, color="555555")
KPI_VAL_FONT = Font(name="Calibri", size=18, bold=True, color="1B365D")
KPI_FILL = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")

ZEBRA_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB")
)

BORDER_TOTAL = Border(
    top=Side(style="thin", color="1B365D"),
    bottom=Side(style="double", color="1B365D")
)

# Soft Status Overlays
STATUS_COLOR_MAP = {
    "MATCHED": {"fill": "E6F4EA", "font": "137333"},       # Green
    "VALUE_MISMATCH": {"fill": "FEF7E0", "font": "B06000"}, # Yellow/Orange
    "MISSING_IN_2B": {"fill": "FCE8E6", "font": "C5221F"},   # Red
    "MISSING_IN_BOOKS": {"fill": "F3E8FF", "font": "6C2BD9"},# Purple
    "PARTIAL_MATCH": {"fill": "E8F0FE", "font": "1A73E8"}    # Blue
}

def format_currency_inr(val: float) -> str:
    """Format float value to INR string format."""
    try:
        return f"₹{val:,.2f}"
    except Exception:
        return "₹0.00"

def get_row_styling(issue_type: str) -> Dict[str, Any]:
    """Return PatternFill and Font for a specific status."""
    status = issue_type.upper()
    if status in STATUS_COLOR_MAP:
        config = STATUS_COLOR_MAP[status]
        return {
            "fill": PatternFill(start_color=config["fill"], end_color=config["fill"], fill_type="solid"),
            "font": Font(name="Calibri", size=10, bold=True, color=config["font"])
        }
    return {
        "fill": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        "font": Font(name="Calibri", size=10, color="000000")
    }

def generate_excel_report(summary: Dict[str, Any], matches: List[Dict[str, Any]], mismatches: List[Dict[str, Any]]) -> bytes:
    """
    Generates a premium, production-grade Excel GST Reconciliation Report using openpyxl.
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHEET 1: EXECUTIVE SUMMARY
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Setup Title Block
    ws_summary.merge_cells("A1:E2")
    title_cell = ws_summary["A1"]
    title_cell.value = "GST RECONCILIATION SUMMARY REPORT"
    title_cell.font = SUMMARY_TITLE_FONT
    title_cell.fill = SUMMARY_TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata Block
    meta_style = Font(name="Calibri", size=10, bold=True, color="374151")
    ws_summary["A4"] = "Client Corporate Name:"
    ws_summary["A4"].font = meta_style
    ws_summary["B4"] = "Client Firm" # Default fallback
    
    ws_summary["A5"] = "Filing Period:"
    ws_summary["A5"].font = meta_style
    ws_summary["B5"] = "March 2024"
    
    ws_summary["D4"] = "Filing GSTIN:"
    ws_summary["D4"].font = meta_style
    ws_summary["E4"] = "27AAACT1234A1Z5"
    
    ws_summary["D5"] = "Filing System:"
    ws_summary["D5"].font = meta_style
    ws_summary["E5"] = "CA-OS Intelligence Engine"
    
    # KPI Blocks
    # Total Invoices KPI
    ws_summary.merge_cells("A7:B7")
    ws_summary["A7"] = "TOTAL RECORDS AUDITED"
    ws_summary["A7"].font = KPI_TITLE_FONT
    ws_summary["A7"].alignment = Alignment(horizontal="center")
    ws_summary.merge_cells("A8:B8")
    ws_summary["A8"] = len(matches) + len(mismatches)
    ws_summary["A8"].font = KPI_VAL_FONT
    ws_summary["A8"].alignment = Alignment(horizontal="center")
    
    # ITC Protected KPI
    ws_summary.merge_cells("D7:E7")
    ws_summary["D7"] = "ITC PROTECTED (ESTIMATED)"
    ws_summary["D7"].font = KPI_TITLE_FONT
    ws_summary["D7"].alignment = Alignment(horizontal="center")
    ws_summary.merge_cells("D8:E8")
    sum_matches = sum(m.get("taxable_value", 0) for m in matches)
    ws_summary["D8"] = format_currency_inr(sum_matches * 0.18)
    ws_summary["D8"].font = Font(name="Calibri", size=18, bold=True, color="137333")
    ws_summary["D8"].alignment = Alignment(horizontal="center")
    
    # Border & Fill for KPIs
    for row in range(7, 9):
        for col in [1, 2, 4, 5]:
            cell = ws_summary.cell(row=row, column=col)
            cell.fill = KPI_FILL
            # Add thin borders around the KPI boxes manually
            
    # Reconciliation Details Table
    ws_summary["A10"] = "Audit Category"
    ws_summary["B10"] = "Invoice Count"
    ws_summary["C10"] = "Taxable Value"
    ws_summary["D10"] = "Estimated Tax (18% ITC)"
    ws_summary["E10"] = "Client Risk Profile"
    
    for col in range(1, 6):
        cell = ws_summary.cell(row=10, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left" if col == 1 else "right" if col in [2, 3, 4] else "center")
        
    categories = [
        ("Matched Invoices", "MATCHED", matches),
        ("Missing in GSTR-2B", "MISSING_IN_2B", [m for m in mismatches if m.get("issue") == "MISSING_IN_2B"]),
        ("Missing in Purchase Books", "MISSING_IN_BOOKS", [m for m in mismatches if m.get("issue") == "MISSING_IN_BOOKS"]),
        ("Value Mismatches", "VALUE_MISMATCH", [m for m in mismatches if m.get("issue") == "VALUE_MISMATCH"]),
        ("Partial Matches", "PARTIAL_MATCH", [m for m in mismatches if m.get("issue") == "PARTIAL_MATCH"]),
    ]
    
    current_row = 11
    total_val = 0.0
    total_count = 0
    
    for label, code, data_list in categories:
        count = len(data_list)
        val = sum(float(x.get("taxable_value", 0)) for x in data_list)
        total_count += count
        total_val += val
        
        ws_summary.cell(row=current_row, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws_summary.cell(row=current_row, column=2, value=count).alignment = Alignment(horizontal="right")
        
        val_cell = ws_summary.cell(row=current_row, column=3, value=val)
        val_cell.number_format = '"₹"#,##,##0.00'
        val_cell.alignment = Alignment(horizontal="right")
        
        tax_cell = ws_summary.cell(row=current_row, column=4, value=val * 0.18)
        tax_cell.number_format = '"₹"#,##,##0.00'
        tax_cell.alignment = Alignment(horizontal="right")
        
        risk_cell = ws_summary.cell(row=current_row, column=5)
        # Style risk profile based on Category
        styling = get_row_styling(code)
        risk_cell.value = "HIGH" if code == "MISSING_IN_2B" else "MEDIUM" if code in ["VALUE_MISMATCH", "MISSING_IN_BOOKS"] else "LOW"
        risk_cell.fill = styling["fill"]
        risk_cell.font = styling["font"]
        risk_cell.alignment = Alignment(horizontal="center")
        
        for c in range(1, 6):
            ws_summary.cell(row=current_row, column=c).border = BORDER_THIN
            
        current_row += 1
        
    # Add Totals Row
    ws_summary.cell(row=current_row, column=1, value="Total Processed").font = Font(name="Calibri", size=10, bold=True)
    ws_summary.cell(row=current_row, column=2, value=total_count).font = Font(name="Calibri", size=10, bold=True)
    
    total_val_cell = ws_summary.cell(row=current_row, column=3, value=total_val)
    total_val_cell.font = Font(name="Calibri", size=10, bold=True)
    total_val_cell.number_format = '"₹"#,##,##0.00'
    total_val_cell.alignment = Alignment(horizontal="right")
    
    total_tax_cell = ws_summary.cell(row=current_row, column=4, value=total_val * 0.18)
    total_tax_cell.font = Font(name="Calibri", size=10, bold=True)
    total_tax_cell.number_format = '"₹"#,##,##0.00'
    total_tax_cell.alignment = Alignment(horizontal="right")
    
    ws_summary.cell(row=current_row, column=5, value="—").alignment = Alignment(horizontal="center")
    
    for c in range(1, 6):
        ws_summary.cell(row=current_row, column=c).border = BORDER_TOTAL
        
    # Auto adjust column widths
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ----------------------------------------------------
    # DETAILS SHEETS
    # ----------------------------------------------------
    sheets_config = [
        ("Matched Invoices", matches, "MATCHED"),
        ("Missing in 2B", [m for m in mismatches if m.get("issue") == "MISSING_IN_2B"], "MISSING_IN_2B"),
        ("Missing in Books", [m for m in mismatches if m.get("issue") == "MISSING_IN_BOOKS"], "MISSING_IN_BOOKS"),
        ("Value Mismatches", [m for m in mismatches if m.get("issue") == "VALUE_MISMATCH"], "VALUE_MISMATCH"),
        ("Partial Matches", [m for m in mismatches if m.get("issue") == "PARTIAL_MATCH"], "PARTIAL_MATCH")
    ]
    
    columns = [
        ("GSTIN", 18, Alignment(horizontal="center")),
        ("Invoice Number", 18, Alignment(horizontal="left")),
        ("Taxable Value", 16, Alignment(horizontal="right")),
        ("Issue Type", 20, Alignment(horizontal="center")),
        ("Likely Cause", 42, Alignment(horizontal="left")),
        ("Recommended Action", 48, Alignment(horizontal="left")),
        ("Risk Level", 14, Alignment(horizontal="center"))
    ]
    
    for sheet_name, data_rows, issue_code in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Sheet Header Style
        ws.cell(row=1, column=1, value=f"CA-OS GST reconciliation working papers — {sheet_name.upper()}").font = Font(name="Calibri", size=12, bold=True, color="1B365D")
        ws.row_dimensions[1].height = 25
        
        # Populate headers in Row 3
        for c_idx, (col_name, width, align) in enumerate(columns, start=1):
            cell = ws.cell(row=3, column=c_idx, value=col_name)
            cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.alignment = align
            ws.column_dimensions[get_column_letter(c_idx)].width = width
            
        ws.row_dimensions[3].height = 24
        
        # Fill data
        d_row = 4
        for r_idx, row_data in enumerate(data_rows):
            fill = ZEBRA_FILL if d_row % 2 == 0 else WHITE_FILL
            
            # Map values
            gstin = row_data.get("gstin") or row_data.get("supplier_gstin") or "—"
            inv_no = row_data.get("invoice_number") or "—"
            tax_val = float(row_data.get("taxable_value", 0))
            
            issue_label = sheet_name if sheet_name != "Matched Invoices" else "Matched"
            likely_cause = row_data.get("likely_cause") or ("N/A - Invoices match perfectly." if sheet_name == "Matched Invoices" else "Unidentified accounting error.")
            rec_action = row_data.get("recommended_action") or ("No action required." if sheet_name == "Matched Invoices" else "Perform a manual visual check.")
            risk = row_data.get("risk_level") or ("LOW" if sheet_name == "Matched Invoices" else "MEDIUM")
            
            # Insert values into cell positions
            c1 = ws.cell(row=d_row, column=1, value=gstin)
            c2 = ws.cell(row=d_row, column=2, value=inv_no)
            
            c3 = ws.cell(row=d_row, column=3, value=tax_val)
            c3.number_format = '"₹"#,##,##0.00'
            
            c4 = ws.cell(row=d_row, column=4, value=issue_label)
            c5 = ws.cell(row=d_row, column=5, value=likely_cause)
            c6 = ws.cell(row=d_row, column=6, value=rec_action)
            
            # Risk styling specifically
            c7 = ws.cell(row=d_row, column=7, value=risk)
            status_style = get_row_styling(issue_code)
            c7.fill = status_style["fill"]
            c7.font = status_style["font"]
            
            # Formatting alignments and fonts
            for c_idx, (col_name, w, align) in enumerate(columns, start=1):
                cell = ws.cell(row=d_row, column=c_idx)
                if c_idx != 7: # Skip overriding risk status fill
                    cell.fill = fill
                    cell.font = Font(name="Calibri", size=10)
                cell.alignment = align
                cell.border = BORDER_THIN
                
            ws.row_dimensions[d_row].height = 20
            d_row += 1
            
        # Empty row checks
        if not data_rows:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
            empty_cell = ws.cell(row=4, column=1, value="No entries found in this audit category.")
            empty_cell.alignment = Alignment(horizontal="center")
            empty_cell.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
            for c_idx in range(1, 8):
                ws.cell(row=4, column=c_idx).border = BORDER_THIN
            ws.row_dimensions[4].height = 25
            
    # Compile Workbook into Memory Bytes
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream.getvalue()


# ----------------------------------------------------
# REPORTLAB PDF CANVAS CALLBACKS (FOR HEADERS/FOOTERS)
# ----------------------------------------------------
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#6B7280"))
        
        # Header (Top of Page)
        self.drawString(36, 756, "CA-OS Chartered Accountant Operating System | Audit Ledger Reports")
        self.setStrokeColor(colors.HexColor("#E5E7EB"))
        self.setLineWidth(0.5)
        self.line(36, 748, 576, 748)
        
        # Footer (Bottom of Page)
        self.line(36, 48, 576, 48)
        self.drawString(36, 36, "CONFIDENTIAL — FOR INTERNAL CA AND CLIENT REVIEW ONLY")
        self.drawRightString(576, 36, f"Page {self._pageNumber} of {page_count}")
        self.restoreState()

def generate_pdf_summary(summary: Dict[str, Any], matches: List[Dict[str, Any]], mismatches: List[Dict[str, Any]]) -> bytes:
    """
    Generates a beautifully styled, high-impact PDF reconciliation summary report using reportlab.
    """
    pdf_stream = BytesIO()
    
    # 0.5 inch margins = 36 points (Letter size is 612 x 792)
    doc = SimpleDocTemplate(
        pdf_stream,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Create Custom Styles (Ensuring Unique Names)
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=22,
        leading=26,
        textColor=colors.HexColor('#1B365D'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#FF7A45'),
        textTransform='uppercase',
        spaceAfter=2
    )
    
    section_h1 = ParagraphStyle(
        'SectionH1',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1B365D'),
        spaceBefore=14,
        spaceAfter=6,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'ReportBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#374151')
    )
    
    bold_body = ParagraphStyle(
        'ReportBoldBody',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    
    meta_label = ParagraphStyle(
        'MetaLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#4B5563')
    )
    
    meta_val = ParagraphStyle(
        'MetaValue',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#111827')
    )
    
    story = []
    
    # ----------------------------------------------------
    # TITLE & BRANDING BLOCK
    # ----------------------------------------------------
    story.append(Paragraph("Intelligence Audit Systems", subtitle_style))
    story.append(Paragraph("GST Reconciliation Report", title_style))
    story.append(Spacer(1, 10))
    
    # ----------------------------------------------------
    # METADATA BLOCK GRID
    # ----------------------------------------------------
    meta_data = [
        [
            Paragraph("CLIENT CORPORATE IDENTITY", meta_label),
            Paragraph("FILING PERIOD MONTH", meta_label),
            Paragraph("AUDITOR PLATFORM SYSTEM", meta_label)
        ],
        [
            Paragraph("Client Firm", meta_val),
            Paragraph("March 2024 (FY 2023-24)", meta_val),
            Paragraph("CA-OS Intelligence Core", meta_val)
        ],
        [
            Paragraph("CLIENT GSTIN ACCOUNT", meta_label),
            Paragraph("REPORT STATUS", meta_label),
            Paragraph("COMPLIANCE VERIFICATION", meta_label)
        ],
        [
            Paragraph("27AAACT1234A1Z5", meta_val),
            Paragraph("ISSUES DETECTED - NEEDS CA REVIEW", ParagraphStyle('RedVal', parent=meta_val, textColor=colors.HexColor('#C5221F'))),
            Paragraph("PRE-AUDIT WORKING PAPERS", meta_val)
        ]
    ]
    
    meta_table = Table(meta_data, colWidths=[200, 180, 160])
    meta_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,1), (-1,1), 0.5, colors.HexColor('#F3F4F6')),
        ('BOTTOMPADDING', (0,1), (-1,1), 6),
        ('TOPPADDING', (0,2), (-1,2), 6),
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 15))
    
    # ----------------------------------------------------
    # SUMMARY METRICS GRID CARDS
    # ----------------------------------------------------
    total_records = len(matches) + len(mismatches)
    exposed_val = sum(float(m.get("taxable_value", 0)) for m in mismatches if m.get("issue") in ["MISSING_IN_2B", "VALUE_MISMATCH"])
    protected_val = sum(float(m.get("taxable_value", 0)) for m in matches)
    
    kpi_data = [
        [
            Paragraph("TOTAL RECORDS AUDITED", ParagraphStyle('KPILbl', parent=meta_label, alignment=1)),
            Paragraph("FULLY RECONCILED (ITC SAFE)", ParagraphStyle('KPILbl', parent=meta_label, alignment=1)),
            Paragraph("MISMATCHES EXPOSED (ITC AT RISK)", ParagraphStyle('KPILbl', parent=meta_label, alignment=1))
        ],
        [
            Paragraph(f"{total_records}", ParagraphStyle('KPIVal', parent=title_style, alignment=1, fontSize=18)),
            Paragraph(f"{len(matches)}", ParagraphStyle('KPIValGreen', parent=title_style, alignment=1, fontSize=18, textColor=colors.HexColor('#137333'))),
            Paragraph(f"{len(mismatches)}", ParagraphStyle('KPIValRed', parent=title_style, alignment=1, fontSize=18, textColor=colors.HexColor('#C5221F')))
        ],
        [
            Paragraph("AGGREGATE INVOICES PROCESS", ParagraphStyle('KPISub', parent=meta_label, alignment=1, fontSize=7, textColor=colors.HexColor('#6B7280'))),
            Paragraph(f"Protected: {format_currency_inr(protected_val * 0.18)}", ParagraphStyle('KPISub', parent=meta_label, alignment=1, fontSize=7, textColor=colors.HexColor('#6B7280'))),
            Paragraph(f"At Risk: {format_currency_inr(exposed_val * 0.18)}", ParagraphStyle('KPISub', parent=meta_label, alignment=1, fontSize=7, textColor=colors.HexColor('#6B7280')))
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[180, 180, 180])
    kpi_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,0), (-1,0), 2),
        ('TOPPADDING', (0,1), (-1,1), 0),
        ('BOTTOMPADDING', (0,1), (-1,1), 0),
        ('TOPPADDING', (0,2), (-1,2), 0),
        ('BOTTOMPADDING', (0,2), (-1,2), 6),
    ]))
    
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # ----------------------------------------------------
    # AUDIT CATEGORY DISTRIBUTION TABLE
    # ----------------------------------------------------
    story.append(Paragraph("Audit Category Distribution Summary", section_h1))
    
    cats = [
        ("Matched Invoices", len(matches), protected_val, "LOW", colors.HexColor('#137333'), colors.HexColor('#E6F4EA')),
        ("Missing in GSTR-2B", len([m for m in mismatches if m.get("issue") == "MISSING_IN_2B"]), sum(float(m.get("taxable_value", 0)) for m in mismatches if m.get("issue") == "MISSING_IN_2B"), "HIGH", colors.HexColor('#C5221F'), colors.HexColor('#FCE8E6')),
        ("Missing in Purchase Books", len([m for m in mismatches if m.get("issue") == "MISSING_IN_BOOKS"]), sum(float(m.get("taxable_value", 0)) for m in mismatches if m.get("issue") == "MISSING_IN_BOOKS"), "MEDIUM", colors.HexColor('#B06000'), colors.HexColor('#FEF7E0')),
        ("Value Mismatches", len([m for m in mismatches if m.get("issue") == "VALUE_MISMATCH"]), sum(float(m.get("taxable_value", 0)) for m in mismatches if m.get("issue") == "VALUE_MISMATCH"), "MEDIUM", colors.HexColor('#B06000'), colors.HexColor('#FEF7E0')),
        ("Partial Matches", len([m for m in mismatches if m.get("issue") == "PARTIAL_MATCH"]), sum(float(m.get("taxable_value", 0)) for m in mismatches if m.get("issue") == "PARTIAL_MATCH"), "LOW", colors.HexColor('#1A73E8'), colors.HexColor('#E8F0FE'))
    ]
    
    cat_header = [
        Paragraph("Reconciliation Audit Category", ParagraphStyle('HCol', parent=bold_body, textColor=colors.white)),
        Paragraph("Count", ParagraphStyle('HColR', parent=bold_body, textColor=colors.white, alignment=2)),
        Paragraph("Taxable Value", ParagraphStyle('HColR', parent=bold_body, textColor=colors.white, alignment=2)),
        Paragraph("ITC Tax (18%)", ParagraphStyle('HColR', parent=bold_body, textColor=colors.white, alignment=2)),
        Paragraph("Risk Profile", ParagraphStyle('HColC', parent=bold_body, textColor=colors.white, alignment=1))
    ]
    
    cat_table_rows = [cat_header]
    
    for idx, (label, cnt, t_val, r_lvl, f_color, bg_color) in enumerate(cats):
        row_style = ParagraphStyle(f'Row{idx}', parent=body_style)
        bold_row = ParagraphStyle(f'BRow{idx}', parent=bold_body)
        
        status_para = Paragraph(f"<font color='{f_color.hexval()}'><b>{r_lvl}</b></font>", ParagraphStyle(f'RCell{idx}', parent=bold_row, alignment=1))
        
        row = [
            Paragraph(f"<b>{label}</b>", row_style),
            Paragraph(f"{cnt}", ParagraphStyle(f'Cnt{idx}', parent=row_style, alignment=2)),
            Paragraph(format_currency_inr(t_val), ParagraphStyle(f'Val{idx}', parent=row_style, alignment=2)),
            Paragraph(format_currency_inr(t_val * 0.18), ParagraphStyle(f'Tax{idx}', parent=row_style, alignment=2)),
            status_para
        ]
        cat_table_rows.append(row)
        
    cat_table = Table(cat_table_rows, colWidths=[180, 60, 110, 110, 80])
    cat_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
    ]))
    
    # Dynamically inject background colors for Risk level cells specifically
    for idx, (label, cnt, t_val, r_lvl, f_color, bg_color) in enumerate(cats, start=1):
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (4, idx), (4, idx), bg_color)
        ]))
        if idx % 2 == 0:
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, idx), (3, idx), colors.HexColor('#F9FAFB'))
            ]))
            
    story.append(cat_table)
    story.append(Spacer(1, 10))
    
    # ----------------------------------------------------
    # TOP CRITICAL MISMATCHES SECTION
    # ----------------------------------------------------
    # Filter for critical ones (High or Medium risk, Missing in 2B or Value mismatch), sort by value descending
    crit_list = [m for m in mismatches if m.get("issue") in ["MISSING_IN_2B", "VALUE_MISMATCH"]]
    crit_sorted = sorted(crit_list, key=lambda x: float(x.get("taxable_value", 0)), reverse=True)[:5]
    
    if crit_sorted:
        story.append(Paragraph("Top Critical Audit Mismatches (By Exposed Value)", section_h1))
        
        crit_header = [
            Paragraph("Invoice Number", ParagraphStyle('CH1', parent=bold_body, textColor=colors.white)),
            Paragraph("Supplier GSTIN", ParagraphStyle('CH2', parent=bold_body, textColor=colors.white)),
            Paragraph("Taxable Value", ParagraphStyle('CH3', parent=bold_body, textColor=colors.white, alignment=2)),
            Paragraph("Exposed ITC", ParagraphStyle('CH4', parent=bold_body, textColor=colors.white, alignment=2)),
            Paragraph("Issue Category", ParagraphStyle('CH5', parent=bold_body, textColor=colors.white, alignment=1))
        ]
        
        crit_rows = [crit_header]
        
        for idx, m in enumerate(crit_sorted, start=1):
            val = float(m.get("taxable_value", 0))
            issue = m.get("issue")
            bg_color = colors.HexColor('#FCE8E6') if issue == "MISSING_IN_2B" else colors.HexColor('#FEF7E0')
            f_color = colors.HexColor('#C5221F') if issue == "MISSING_IN_2B" else colors.HexColor('#B06000')
            
            row = [
                Paragraph(m.get("invoice_number", "—"), body_style),
                Paragraph(m.get("gstin") or m.get("supplier_gstin") or "—", ParagraphStyle(f'Gst{idx}', parent=body_style, fontName='Helvetica-Bold')),
                Paragraph(format_currency_inr(val), ParagraphStyle(f'CritV{idx}', parent=body_style, alignment=2)),
                Paragraph(format_currency_inr(val * 0.18), ParagraphStyle(f'CritT{idx}', parent=body_style, alignment=2)),
                Paragraph(f"<b>{issue.replace('_', ' ')}</b>", ParagraphStyle(f'CritI{idx}', parent=body_style, alignment=1, textColor=f_color))
            ]
            crit_rows.append(row)
            
        crit_table = Table(crit_rows, colWidths=[100, 130, 110, 110, 90])
        crit_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ]))
        
        for idx, m in enumerate(crit_sorted, start=1):
            issue = m.get("issue")
            bg = colors.HexColor('#FCE8E6') if issue == "MISSING_IN_2B" else colors.HexColor('#FEF7E0')
            crit_table.setStyle(TableStyle([
                ('BACKGROUND', (4, idx), (4, idx), bg)
            ]))
            if idx % 2 == 0:
                crit_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, idx), (3, idx), colors.HexColor('#F9FAFB'))
                ]))
                
        story.append(crit_table)
        story.append(Spacer(1, 15))
        
    # ----------------------------------------------------
    # AUDITOR DECLARATION & SIGNATURE (KEEP TOGETHER)
    # ----------------------------------------------------
    sig_data = [
        [
            Paragraph("<b>CA AUDIT & PROFESSIONAL COMPLIANCE STATEMENT</b><br/>"
                      "This report presents reconciliation pre-audit working papers evaluated automatically in-memory. "
                      "All findings (including missing invoices and taxable mismatches) should be verified manually "
                      "by the principal auditor against physical invoices prior to finalizing corporate tax filings.", ParagraphStyle('SigD', parent=body_style, fontSize=7, leading=10, textColor=colors.HexColor('#4B5563'))),
            Paragraph("<b>VERIFIED AUDITOR SEAL</b><br/><br/>"
                      "Principal Signatory: ________________________<br/>"
                      "Reckon CA Firm Partnership", ParagraphStyle('SigS', parent=body_style, fontSize=7, leading=10))
        ]
    ]
    
    sig_table = Table(sig_data, colWidths=[360, 180])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('LINEAFTER', (0,0), (0,0), 0.5, colors.HexColor('#E2E8F0'))
    ]))
    
    # Wrap in KeepTogether to ensure it doesn't break across pages awkwardly
    story.append(KeepTogether([sig_table]))
    
    # Build Document using Custom Page Numbered Canvas
    doc.build(story, canvasmaker=NumberedCanvas)
    
    pdf_stream.seek(0)
    return pdf_stream.getvalue()
